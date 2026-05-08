"""
XlsxService — importação e sincronização de clientes via planilha Excel.

Regras:
  - CNPJ é a chave primária de sincronização
  - UPDATE  se o CNPJ já existir no repositório
  - INSERT  se o CNPJ não existir
  - DELETE  se o CNPJ existir no repositório mas não estiver na nova planilha
"""
import re
import uuid
from dataclasses import dataclass
from typing import Optional

from app.models.client import Client


@dataclass
class SyncResult:
    inserted: list[Client]
    updated: list[Client]
    deleted: list[Client]
    errors: list[str]

    @property
    def total(self):
        return len(self.inserted) + len(self.updated) + len(self.deleted)

    def summary(self) -> str:
        lines = [
            f"✅ Inseridos : {len(self.inserted)}",
            f"🔄 Atualizados: {len(self.updated)}",
            f"🗑 Removidos : {len(self.deleted)}",
        ]
        if self.errors:
            lines.append(f"⚠ Erros     : {len(self.errors)}")
            for e in self.errors:
                lines.append(f"   • {e}")
        return "\n".join(lines)


class XlsxService:
    """Lê planilha Excel e sincroniza com o repositório."""

    COLUNAS_ESPERADAS = {"NOME", "CODIGO", "CNPJ"}

    def load_from_file(self, path: str) -> tuple[list[Client], list[str]]:
        """
        Lê o arquivo .xlsx e retorna (lista de clientes, lista de erros).
        Não toca no repositório — só lê e valida.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], ["openpyxl não instalado. Execute: pip install openpyxl"]

        try:
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            return [], [f"Erro ao abrir planilha: {e}"]

        if not rows:
            return [], ["Planilha vazia."]

        # Detecta cabeçalho
        header = [str(c).strip().upper() if c else "" for c in rows[0]]
        missing = self.COLUNAS_ESPERADAS - set(header)
        if missing:
            return [], [f"Colunas obrigatórias não encontradas: {', '.join(missing)}"]

        idx = {col: header.index(col) for col in header if col}
        nome_idx   = idx.get("NOME")
        codigo_idx = idx.get("CODIGO")
        cnpj_idx   = idx.get("CNPJ")
        ie_idx     = idx.get("IE")

        clients = []
        errors  = []

        for i, row in enumerate(rows[1:], start=2):
            try:
                nome   = str(row[nome_idx]).strip()   if row[nome_idx]   else ""
                codigo = str(row[codigo_idx]).strip() if row[codigo_idx] else ""
                cnpj   = re.sub(r'\D', '', str(row[cnpj_idx])) if row[cnpj_idx] else ""
                ie_raw = str(row[ie_idx]).strip()     if (ie_idx is not None and row[ie_idx]) else None

                if not nome:
                    errors.append(f"Linha {i}: nome vazio — ignorada.")
                    continue
                if not cnpj or len(cnpj) < 11:
                    errors.append(f"Linha {i} ({nome}): CNPJ inválido '{cnpj}' — ignorada.")
                    continue

                # IE: ignora string vazia
                ie = ie_raw if (ie_raw and ie_raw not in ("", "None")) else None

                clients.append(Client(
                    id=str(uuid.uuid5(uuid.NAMESPACE_DNS, cnpj)),  # ID determinístico por CNPJ
                    company_name=nome,
                    cnpj=cnpj,
                    codigo=codigo,
                    inscricao_estadual=ie,
                ))
            except Exception as e:
                errors.append(f"Linha {i}: erro inesperado — {e}")

        return clients, errors

    def sync(self, path: str, repository) -> SyncResult:
        """
        Carrega planilha e aplica INSERT / UPDATE / DELETE no repositório.
        """
        novos_clientes, errors = self.load_from_file(path)
        if not novos_clientes and errors:
            return SyncResult(inserted=[], updated=[], deleted=[], errors=errors)

        existentes = {c.cnpj_digits(): c for c in repository.list_clients(active_only=False)}
        novos_map  = {c.cnpj_digits(): c for c in novos_clientes}

        inserted = []
        updated  = []
        deleted  = []

        # INSERT e UPDATE
        for cnpj_d, novo in novos_map.items():
            if cnpj_d in existentes:
                # Preserva campos que a planilha não tem (email, responsible, etc.)
                antigo = existentes[cnpj_d]
                novo.id          = antigo.id
                novo.email       = antigo.email
                novo.responsible = antigo.responsible
                novo.department  = antigo.department
                novo.phone       = antigo.phone
                repository.save_client(novo)
                updated.append(novo)
            else:
                repository.save_client(novo)
                inserted.append(novo)

        # DELETE — clientes no repositório que não vieram na planilha
        for cnpj_d, existente in existentes.items():
            if cnpj_d not in novos_map:
                repository.delete_client(existente.id)
                deleted.append(existente)

        return SyncResult(inserted=inserted, updated=updated, deleted=deleted, errors=errors)
